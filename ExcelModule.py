import numpy as np
import pandas as pd
from openpyxl import *
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Series, Reference
from datetime import datetime, timedelta
import sys

sys.path.insert(0, "C:\\Users\\dmanassi.ACROSS\\OneDrive - Ediscom Spa\\Desktop\\Moduli")
import Colors as clr

def cella_titolo(ws, row, col, style, color):
    ws.cell(row, col).fill = PatternFill('solid', start_color=color, end_color=color)
    ws.cell(row, col).font = Font(bold=True)
    ws.cell(row, col).border = Border(left=Side(style=style), right=Side(style=style),
                                      bottom=Side(style=style), top=Side(style=style))
    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def cella(ws, row, col, style):
    ws.cell(row, col).border = Border(left=Side(style=style), right=Side(style=style),
                                      bottom=Side(style=style), top=Side(style=style))
    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')

def rate(value1, value2):
    if value1 != 0:
        try:
            result = value1 / value2
        except ZeroDivisionError:
            result = 0
    else:
        result = 0

    return result

def verifica_soglia(ws, val1, val2, riga_arrivo, colonna_arrivo):
    if val1 >= val2:
        ws.cell(riga_arrivo, colonna_arrivo).value = "SI"
        ws.cell(riga_arrivo, colonna_arrivo).fill = PatternFill('solid', start_color=clr.verde, end_color=clr.verde)
    else:
        ws.cell(riga_arrivo, colonna_arrivo).value = "NO"
        ws.cell(riga_arrivo, colonna_arrivo).fill = PatternFill('solid', start_color=clr.rosso, end_color=clr.rosso)
